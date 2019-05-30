from tkinter import *
import openpyxl

wb = openpyxl.load_workbook('database.xlsx')
ws = wb['Pharmacy']
current_row = 1
last_row = ws.max_row + 1


def add_item():
    global last_row
    ws.cell(last_row, 1).value = item_name.get()
    ws.cell(last_row, 2).value = item_price.get()
    ws.cell(last_row, 3).value = item_quant.get()
    ws.cell(last_row, 4).value = item_categ.get()
    ws.cell(last_row, 5).value = item_disc.get()
    clear_fields()


def view_first():
    global current_row
    current_row = 2
    item_name.set(ws.cell(2, 1).value)
    item_price.set(ws.cell(2, 2).value)
    item_quant.set(ws.cell(2, 3).value)
    item_categ.set(ws.cell(2, 4).value)
    item_disc.set(ws.cell(2, 5).value)


def view_last():
    global current_row
    current_row = ws.max_row
    item_name.set(ws.cell(ws.max_row, 1).value)
    item_price.set(ws.cell(ws.max_row, 2).value)
    item_quant.set(ws.cell(ws.max_row, 3).value)
    item_categ.set(ws.cell(ws.max_row, 4).value)
    item_disc.set(ws.cell(ws.max_row, 5).value)


def view_next():
    global current_row
    current_row += 1
    item_name.set(ws.cell(current_row, 1).value)
    item_price.set(ws.cell(current_row, 2).value)
    item_quant.set(ws.cell(current_row, 3).value)
    item_categ.set(ws.cell(current_row, 4).value)
    item_disc.set(ws.cell(current_row, 5).value)


def view_previous():
    global current_row
    current_row -= 1
    item_name.set(ws.cell(current_row, 1).value)
    item_price.set(ws.cell(current_row, 2).value)
    item_quant.set(ws.cell(current_row, 3).value)
    item_categ.set(ws.cell(current_row, 4).value)
    item_disc.set(ws.cell(current_row, 5).value)


def update_item():
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == item_name.get():
                ws.cell(cell.row, 2).value = item_price.get()
                ws.cell(cell.row, 3).value = item_quant.get()
                ws.cell(cell.row, 4).value = item_categ.get()
                ws.cell(cell.row, 5).value = item_disc.get()
    clear_fields()


def delete_item():
    global last_row
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == item_name.get():
                ws.delete_rows(cell.row, 1)
    last_row -= 1
    clear_fields()


def search_item():
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == item_name.get():
                item_price.set(ws.cell(cell.row, 2).value)
                item_quant.set(ws.cell(cell.row, 3).value)
                item_categ.set(ws.cell(cell.row, 4).value)
                item_disc.set(ws.cell(cell.row, 5).value)


def clear_fields():
    name_entry.delete(0, END)
    price_entry.delete(0, END)
    quant_entry.delete(0, END)
    categ_entry.delete(0, END)
    disc_entry.delete(0, END)


root = Tk()
root.title('Pharmacy Management System')
root.geometry('1000x500')
root.configure(background='black')
root.resizable(False, False)

# Variables
item_name = StringVar()
item_price = StringVar()
item_quant = StringVar()
item_categ = StringVar()
item_disc = StringVar()

# Labels
title_label = Label(root, text='PHARMACY MANAGEMENT SYSTEM', bg='black', fg='white', font='none 20 bold')
title_label.place(x=260, y=35)
name_label = Label(root, text='ENTER ITEM NAME', bg='red', fg='white', font='none 10', width=25)
name_label.place(x=80, y=100)
price_label = Label(root, text='ENTER ITEM PRICE', bg='red', fg='white', font='none 10', width=25)
price_label.place(x=80, y=160)
quant_label = Label(root, text='ENTER ITEM QUANTITY', bg='red', fg='white', font='none 10', width=25)
quant_label.place(x=80, y=220)
categ_label = Label(root, text='ENTER ITEM CATEGORY', bg='red', fg='white', font='none 10', width=25)
categ_label.place(x=80, y=280)
disc_label = Label(root, text='ENTER ITEM DISCOUNT', bg='red', fg='white', font='none 10', width=25)
disc_label.place(x=80, y=340)

# Entries
name_entry = Entry(root, textvariable=item_name, width=25)
name_entry.place(x=320, y=100)
price_entry = Entry(root, textvariable=item_price, width=25)
price_entry.place(x=320, y=160)
quant_entry = Entry(root, textvariable=item_quant, width=25)
quant_entry.place(x=320, y=220)
categ_entry = Entry(root, textvariable=item_categ, width=25)
categ_entry.place(x=320, y=280)
disc_entry = Entry(root, textvariable=item_disc, width=25)
disc_entry.place(x=320, y=340)

# Buttons
add_btn = Button(root, text='ADD ITEM', font='none 10', width=20, command=add_item)
add_btn.place(x=520, y=100)
first_btn = Button(root, text='VIEW FIRST ITEM', font='none 10', width=20, command=view_first)
first_btn.place(x=520, y=160)
previous_btn = Button(root, text='VIEW PREVIOUS ITEM', font='none 10', width=20, command=view_previous)
previous_btn.place(x=520, y=220)
update_btn = Button(root, text='UPDATE ITEM', font='none 10', width=20, command=update_item)
update_btn.place(x=520, y=280)
del_btn = Button(root, text='DELETE ITEM', font='none 10', width=20, command=delete_item)
del_btn.place(x=750, y=100)
next_btn = Button(root, text='VIEW NEXT ITEM', font='none 10', width=20, command=view_next)
next_btn.place(x=750, y=160)
last_btn = Button(root, text='VIEW LAST ITEM', font='none 10', width=20, command=view_last)
last_btn.place(x=750, y=220)
search_btn = Button(root, text='SEARCH ITEM', font='none 10', width=20, command=search_item)
search_btn.place(x=750, y=280)
clear_btn = Button(root, text='CLEAR SCREEN', font='none 10', width=20, command=clear_fields)
clear_btn.place(x=750, y=340)

root.mainloop()
wb.save('database.xlsx')
